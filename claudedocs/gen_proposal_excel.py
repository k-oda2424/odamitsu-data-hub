# -*- coding: utf-8 -*-
"""メイプルシティ提案見積書 Excel生成
欄外（I-J列）に粗利率・仕入原価を配置。
粗利率セルを変更するだけで見積単価→年間合計→削減額がすべて自動更新。
"""
import math
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== Styles =====
font_title = Font(name='游ゴシック', size=16, bold=True)
font_subtitle = Font(name='游ゴシック', size=11, bold=True)
font_header = Font(name='游ゴシック', size=9, bold=True, color='FFFFFF')
font_body = Font(name='游ゴシック', size=9)
font_body_bold = Font(name='游ゴシック', size=9, bold=True)
font_star = Font(name='游ゴシック', size=9, bold=True, color='D94600')
font_note = Font(name='游ゴシック', size=8, color='666666')
font_red = Font(name='游ゴシック', size=9, bold=True, color='CC0000')
font_blue = Font(name='游ゴシック', size=9, bold=True, color='0066CC')
font_section = Font(name='游ゴシック', size=10, bold=True)
font_setting_label = Font(name='游ゴシック', size=10, bold=True, color='2F5496')
font_setting_val = Font(name='游ゴシック', size=14, bold=True, color='CC0000')
font_cost_header = Font(name='游ゴシック', size=8, bold=True, color='888888')
font_cost = Font(name='游ゴシック', size=8, color='888888')

fill_header = PatternFill('solid', fgColor='2F5496')
fill_kao_cur = PatternFill('solid', fgColor='F2F2F2')
fill_kao_new = PatternFill('solid', fgColor='FFF2CC')
fill_alt = PatternFill('solid', fgColor='B4C6E7')
fill_section = PatternFill('solid', fgColor='D9E2F3')
fill_highlight = PatternFill('solid', fgColor='E2EFDA')
fill_total = PatternFill('solid', fgColor='D9E2F3')
fill_setting = PatternFill('solid', fgColor='FFF2CC')

thin = Side(style='thin', color='CCCCCC')
border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
bold_border = Border(
    top=Side(style='double', color='2F5496'),
    bottom=Side(style='double', color='2F5496'),
    left=thin, right=thin
)
border_setting = Border(
    top=Side(style='medium', color='CC0000'),
    bottom=Side(style='medium', color='CC0000'),
    left=Side(style='medium', color='CC0000'),
    right=Side(style='medium', color='CC0000'),
)
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center')
num_yen = '#,##0"円"'

NCOLS = 6  # A-F: No, 商品名, 規格, 花王改定, 代替品, 備考


def sc(ws, row, col, value, font=font_body, fill=None, alignment=align_center,
       border=border_all, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


# ===== Sheet setup =====
ws = wb.active
ws.title = '提案見積書'

# A:No B:商品名 C:規格 D:花王改定 E:代替品 F:備考 | H:gap I:花王原価 J:代替原価
col_widths = {1: 4, 2: 28, 3: 18, 4: 15, 5: 16, 6: 38,
              7: 2, 8: 2, 9: 14, 10: 14, 11: 10, 12: 10, 13: 10, 14: 10}
for i, w in col_widths.items():
    ws.column_dimensions[get_column_letter(i)].width = w

# ===== Settings area (I-J columns, outside print area) =====
MARGIN_CELL = '$J$2'

sc(ws, 1, 9, '【粗利率設定】', font_setting_label, fill_setting, align_left, border=None)
sc(ws, 2, 9, '粗利率', font_setting_label, fill_setting, align_right, border=border_all)
sc(ws, 2, 10, 0.15, font_setting_val, fill_setting, align_center, border=border_setting,
   number_format='0%')
sc(ws, 3, 9, '', border=None)
sc(ws, 3, 10, '↑ この値を変更', font_note, alignment=align_left, border=None)

font_indiv = Font(name='游ゴシック', size=9, bold=True, color='CC0000')
fill_indiv = PatternFill('solid', fgColor='FFF2CC')
border_indiv = Border(
    top=Side(style='thin', color='CC0000'),
    bottom=Side(style='thin', color='CC0000'),
    left=Side(style='thin', color='CC0000'),
    right=Side(style='thin', color='CC0000'),
)

sc(ws, 5, 9, '花王値上後原価', font_cost_header, alignment=align_center, border=None)
sc(ws, 5, 10, '代替品原価', font_cost_header, alignment=align_center, border=None)
sc(ws, 5, 11, '花王実粗利', font_cost_header, alignment=align_center, border=None)
sc(ws, 5, 12, '代替実粗利', font_cost_header, alignment=align_center, border=None)
sc(ws, 5, 13, '花王個別粗利', font_indiv, fill_indiv, align_center, border=None)
sc(ws, 5, 14, '代替個別粗利', font_indiv, fill_indiv, align_center, border=None)
sc(ws, 6, 13, '※空欄=全体適用', font_note, alignment=align_center, border=None)
sc(ws, 6, 14, '※空欄=全体適用', font_note, alignment=align_center, border=None)

# ===== Title & Meta =====
ws.merge_cells('A1:F1')
c = ws['A1']
c.value = 'メイプルシティ様 お見積書'
c.font = font_title
c.alignment = Alignment(horizontal='center', vertical='center')

meta = [
    ('日付', '2026年4月'),
    ('件名', '花王プロフェッショナル製品 価格改定のご案内 および 代替商品のご提案'),
    ('適用', '2026年5月1日〜'),
]
for i, (k, v) in enumerate(meta, 3):
    sc(ws, i, 1, k, font_body_bold, border=None, alignment=align_left)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=NCOLS)
    sc(ws, i, 2, v, font_body, border=None, alignment=align_left)

ws.merge_cells('A7:F8')
c = ws['A7']
c.value = (
    '花王プロフェッショナルサービスより2026年5月1日出荷分より仕入価格の改定が実施されます。\n'
    'あわせて、コスト削減にお役立ていただけるライオンハイジーン製品・セッツ製品もご提案いたします。'
)
c.font = font_body
c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ===== Headers =====
headers = ['No', '商品名', '規格', '花王改定価格\n（税抜）',
           '代替品 見積単価\n（税抜）', '備考']
row = 10
for col, h in enumerate(headers, 1):
    sc(ws, row, col, h, font_header, fill_header, align_center)

# ===== Product data =====
# (no, category, kao_name, kao_spec, kao_cur_sell, kao_old_cost, kao_new_cost,
#  alt_name, alt_spec, alt_cost, alt_mult, note, is_star)
products = [
    (1, '食器用洗剤', 'ファミリーフレッシュ', '4.5L（原液）', 1400, 1238, 1423,
     '★ ダイバークリーンコンク【セッツ】', '5Kg（6倍希釈=実質30kg）', 1650, 1,
     '6倍希釈で実質30kg分。洗浄力2.5倍。ランニングコスト約1/6', True),
    (2, '食器用洗剤（濃縮）', 'パワーストリーム コンク', '5L', 2807, 2514, 2891,
     '★ セルシアコンクα【ライオン】', '4L', 2120, 1,
     '同等6倍希釈濃縮タイプ', True),
    (3, '漂白剤', 'ハイターE', '5Kg', 780, 618, 680,
     '★ キッチンブリーチ D-104【セッツ】', '5Kg', 520, 1,
     '同等塩素系。同規格で直接置換可', True),
    (4, '厨房油汚れ洗剤', 'パワークリーナー（パワーマジ）', '4.5L', 1880, 1644, 1809,
     '★ グリースサットル【ライオン】', '5Kg', 1250, 1,
     '除菌+泡密着タイプ', True),
    (5, '洗濯用洗剤', 'アタック 消臭ストロングジェル', '4kg', 2020, 1358, 1494,
     '★ トップクリアリキッド【ライオン】', '4Kg', 1400, 1,
     '抗菌・ウイルス除去', True),
    (6, '手洗い石けん', 'C&C F1 薬用泡ハンドウォッシュ', '4L', 3600, 2936, 3230,
     '★ キレイキレイ薬用泡ﾊﾝﾄﾞｿｰﾌﾟ プロ無香料【ライオン】', '4L', 1850, 1,
     '医薬部外品・食品施設向け無香料', True),
    (7, '手指消毒剤', 'ハンドスキッシュEX 詰替', '4.5L', 3746, 2391, 2630,
     '★ サニテートAハンドミスト【ライオン】', '4L', 2150, 1,
     '64vol%エタノール・非危険物', True),
    (8, 'アルコール製剤', 'スキッシュV', '4.5L', 2300, 2129, 2341,
     '★ ユービコールノロV【セッツ】', '5L', 1980, 1,
     'ノロウイルス対応！ 軽減税率8%・容量11%増', True),
    (9, 'アルコール製剤', 'スキッシュV', '10L', 4950, 4509, 4960,
     '★ ユービコールノロV66【セッツ】×2本', '5L×2', 2200, 2,
     'ノロウイルス対応(高濃度66vol%)・軽減税率8%', True),
    (10, 'トイレ用洗剤', 'トイレマジックリン 消臭ストロング', '4.5L', 2530, 1980, 2178,
     '★ おそうじルック【ライオン】', '4L', 850, 1,
     '中性・除菌効果あり', True),
    (11, '食器洗浄機用', 'パワークリンキーパー 高速すすぎ', '5Kg', 3191, 2415, 2656,
     '（花王継続）', '—', None, 1,
     '代替品は別途ご相談', False),
    (12, 'スチコン洗浄剤', 'スチコンクリーナー', '2L', 1320, 1054, 1160,
     '（花王継続）', '—', None, 1,
     '花王継続推奨', False),
    (13, '除菌シート', 'セイフキープEX ﾜｲﾄﾞﾌﾀ付ﾋﾟﾛｰ', '80枚', 650, 505, 556,
     '★ メディプロ エタノールクロス【ライオン】', '80枚', 330, 1,
     'エタノール含浸除菌シート', True),
]

alt_price_rows = {}  # product_no -> alt_row
kao_price_rows = {}  # product_no -> kao_row

row = 11
for p in products:
    (no, cat, kao_name, kao_spec, kao_cur_sell, kao_old_cost, kao_new_cost,
     alt_name, alt_spec, alt_cost, alt_mult, note, is_star) = p

    # Category row
    for c in range(1, NCOLS + 1):
        sc(ws, row, c, '', font_section, fill_section, align_left)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    ws.cell(row=row, column=1).value = f'  {cat}'
    row += 1

    # Kao row
    kao_price_rows[no] = row
    sc(ws, row, 1, no, font_body, fill_kao_cur, align_center)
    sc(ws, row, 2, f'花王 {kao_name}', font_body, fill_kao_cur, align_left)
    sc(ws, row, 3, kao_spec, font_body, fill_kao_cur, align_center)
    # D: 花王改定 = CEILING(原価*(1+粗利率), 5) — 個別があれば個別優先
    # IF(M{row}<>"", M{row}, $J$2) で適用粗利を決定
    kao_margin = f'IF(M{row}<>"",M{row},{MARGIN_CELL})'
    sc(ws, row, 4, f'=CEILING(I{row}*(1+{kao_margin}),5)', font_body_bold,
       fill_kao_new, align_right, number_format=num_yen)

    # Settings area: 花王値上後原価 + 実粗利率 + 個別粗利入力欄
    sc(ws, row, 9, kao_new_cost, font_cost, alignment=align_right, border=None, number_format='#,##0')
    sc(ws, row, 11, f'=(D{row}-I{row})/I{row}', font_cost, alignment=align_right,
       border=None, number_format='0.0%')
    # M列: 花王個別粗利（空欄=全体適用）
    sc(ws, row, 13, None, font_indiv, fill_indiv, align_center, border=border_indiv,
       number_format='0%')

    if alt_cost is not None:
        sc(ws, row, 5, '', font_body, fill_kao_cur)
        sc(ws, row, 6, f'現行 {kao_cur_sell:,}円', font_note, fill_kao_cur, align_left)
    else:
        sc(ws, row, 5, '—', font_body, fill_kao_cur, align_center)
        sc(ws, row, 6, f'現行 {kao_cur_sell:,}円 / {note}', font_note, fill_kao_cur, align_left)
    row += 1

    # Alt row with formula
    if alt_cost is not None:
        alt_row = row
        alt_price_rows[no] = alt_row

        f_name = font_star if is_star else font_body
        sc(ws, row, 1, '', font_body, fill_highlight)
        sc(ws, row, 2, alt_name, f_name, fill_highlight, align_left)
        sc(ws, row, 3, alt_spec, font_body, fill_highlight, align_center)
        sc(ws, row, 4, '', font_body, fill_highlight)

        # Settings area: 代替品原価 + 実粗利率 + 個別粗利入力欄
        sc(ws, row, 10, alt_cost, font_cost, alignment=align_right, border=None, number_format='#,##0')

        # E: =CEILING(原価*(1+粗利率),5) [*mult] — 個別があれば個別優先
        cost_ref = f'J{row}'
        alt_margin = f'IF(N{row}<>"",N{row},{MARGIN_CELL})'
        if alt_mult > 1:
            formula = f'=CEILING({cost_ref}*(1+{alt_margin}),5)*{alt_mult}'
            margin_formula = f'=(E{row}/{alt_mult}-J{row})/J{row}'
        else:
            formula = f'=CEILING({cost_ref}*(1+{alt_margin}),5)'
            margin_formula = f'=(E{row}-J{row})/J{row}'

        sc(ws, row, 5, formula, font_blue, fill_alt, align_right, number_format=num_yen)
        sc(ws, row, 6, note, font_note, fill_highlight, align_left)
        sc(ws, row, 12, margin_formula, font_cost, alignment=align_right,
           border=None, number_format='0.0%')
        # N列: 代替品個別粗利（空欄=全体適用）
        sc(ws, row, 14, None, font_indiv, fill_indiv, align_center, border=border_indiv,
           number_format='0%')
        row += 1

row += 1

# ===== Annual comparison =====
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
sc(ws, row, 1, '年間コスト比較（税抜）', font_subtitle, alignment=align_left, border=None)
row += 1

ann_headers = ['', '品目 → 代替品', '年間数量', '花王改定 年間',
               '代替品 年間', '削減額']
for col, h in enumerate(ann_headers, 1):
    sc(ws, row, col, h, font_header, fill_header, align_center)
row += 1

diver_bottles = math.ceil(161 * 4500 / 30000)

annual_items = [
    ('食器洗剤→ダイバークリーンコンク', '161→25本', 161, diver_bottles, 1),
    ('漂白剤→セッツキッチンブリーチ', '232本', 232, 232, 3),
    ('油汚れ→グリースサットル', '91本', 91, 91, 4),
    ('洗濯→トップクリアリキッド', '33本', 33, 33, 5),
    ('ハンドソープ→キレイキレイプロ', '21本', 21, 21, 6),
    ('手指消毒→サニテートA', '24本', 24, 24, 7),
    ('アルコール4.5L→ノロV 5L', '4本', 4, 4, 8),
    ('アルコール10L→ノロV66 5L×2', '9本', 9, 9, 9),
    ('トイレ→おそうじルック', '4本', 4, 4, 10),
    ('濃縮台所→セルシアコンクα', '8本', 8, 8, 2),
    ('除菌シート→メディプロ', '44個', 44, 44, 13),
    ('食洗機（花王継続）', '102本', 102, 102, 11),
    ('スチコン（花王継続）', '3本', 3, 3, 12),
]

ann_rows = []
for label, qty_str, kao_ann_qty, alt_ann_qty, prod_no in annual_items:
    sc(ws, row, 1, '', font_body)
    sc(ws, row, 2, label, font_body, alignment=align_left)
    sc(ws, row, 3, qty_str, font_body, alignment=align_center)

    kao_row = kao_price_rows[prod_no]

    # D: 花王改定年間 = 数量 × D{kao_row}
    if prod_no == 1:
        sc(ws, row, 4, f'=161*D{kao_row}', font_body,
           alignment=align_right, number_format='#,##0')
    else:
        sc(ws, row, 4, f'={kao_ann_qty}*D{kao_row}', font_body,
           alignment=align_right, number_format='#,##0')

    # E: 代替品年間
    if prod_no in alt_price_rows:
        alt_row = alt_price_rows[prod_no]
        sc(ws, row, 5, f'={alt_ann_qty}*E{alt_row}', font_blue,
           alignment=align_right, number_format='#,##0')
    else:
        # 花王継続: 代替=花王改定
        sc(ws, row, 5, f'=D{row}', font_body,
           alignment=align_right, number_format='#,##0')

    # F: 削減額 = 代替 - 花王改定
    sc(ws, row, 6, f'=E{row}-D{row}', font_blue,
       alignment=align_right, number_format='#,##0')

    ann_rows.append(row)
    row += 1

# Totals
first = ann_rows[0]
last = ann_rows[-1]
sc(ws, row, 1, '', font_body_bold, fill_total, border=bold_border)
sc(ws, row, 2, '合計', font_body_bold, fill_total, align_left, bold_border)
sc(ws, row, 3, '', font_body_bold, fill_total, border=bold_border)
for col_i in [4, 5]:
    f = font_blue if col_i == 5 else font_body_bold
    col_l = get_column_letter(col_i)
    sc(ws, row, col_i, f'=SUM({col_l}{first}:{col_l}{last})',
       f, fill_total, align_right, bold_border, '#,##0')
sc(ws, row, 6, f'=E{row}-D{row}', font_red, fill_total, align_right, bold_border, '#,##0')
total_row = row
row += 2

# Notes
notes = [
    '※ 食器洗剤はダイバークリーンコンク（6倍希釈）の同等洗浄量換算（161本→約25本）',
    '※ ノロV/V66は食品添加物のため軽減税率8%適用。税込ベースではさらに差額拡大',
    '※ 花王製品・ライオン製品・セッツ製品とも、商品単位でお選びいただけます',
    '※ ダイバークリーンコンクは800gパウチ、ノロVはトライアルセットからお試し可能',
    '※ すべて税抜価格です',
]
for note in notes:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    sc(ws, row, 1, note, font_note, alignment=align_left, border=None)
    row += 1

row += 1
sc(ws, row, 1, '小田光株式会社', font_body_bold, alignment=align_left, border=None)

# Print settings (A-F only)
ws.print_area = f'A1:F{row}'
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

path = r'C:\project\odamitsu-data-hub\claudedocs\proposal-maple-city-202605.xlsx'
wb.save(path)
print(f'Saved: {path}')
